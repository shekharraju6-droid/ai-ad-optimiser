import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

wb = load_workbook("data/shym_steel_report.xlsx", read_only=True, data_only=True)
with open("data/inspect.txt", "w", encoding="utf-8") as out:
    out.write(f"Sheets: {wb.sheetnames}\n\n")
    for name in wb.sheetnames:
        ws = wb[name]
        out.write(f"--- {name} (max row {ws.max_row}, max col {ws.max_column}) ---\n")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 15:
                out.write("...\n")
                break
            out.write(str(row) + "\n")
        out.write("\n")
wb.close()
print("Inspect written to data/inspect.txt")
