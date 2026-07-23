"""
Mantri MIS Reports endpoints for InsightDesk.

Provides persisted daily snapshots of Google Ads and Meta Ads performance
per MIS project, plus refresh, export (PDF/Excel), and aggregate views.
"""
import io
import logging
from datetime import date, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from backend.db.database import get_db
from backend.db.models import Account, MisProject, MisDailySnapshot, User
from backend.routes.auth import get_current_user_required
from backend.services.activity_log import log_activity
from backend.services.connectors import get_connector

logger = logging.getLogger("AdOptima")

router = APIRouter(prefix="/api/mis/mantri", tags=["mis_mantri"])


def _fmt_inr(n: Optional[float]) -> str:
    if n is None:
        return "₹0"
    return "₹" + "{:,.0f}".format(round(n))


def _ordinal_date(d: date) -> str:
    day = d.day
    if 11 <= day <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(day % 10, "th")
    return f"{day}{suffix} {d.strftime('%b')} {d.year}"


def _default_date_range(days: int = 30) -> tuple:
    end = date.today() - timedelta(days=1)
    start = end - timedelta(days=days - 1)
    return start.isoformat(), end.isoformat()


# Hardcoded platform start dates for Mantri project
GOOGLE_START_DATE = "2026-05-25"
META_START_DATE = "2026-04-17"


def _parse_date(d_str: str) -> date:
    if not d_str:
        raise HTTPException(status_code=400, detail="Date string cannot be empty")
    d_str = str(d_str).strip()
    try:
        parts = d_str.split("-")
        if len(parts) == 3:
            if len(parts[0]) == 2 and len(parts[2]) == 4:
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                res = date(y, m, d)
            elif len(parts[0]) == 4:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                res = date(y, m, d)
            else:
                res = date.fromisoformat(d_str)
        else:
            res = date.fromisoformat(d_str)
        if res.year < 2000 or res.year > 2100:
            raise ValueError(f"Year {res.year} out of realistic range (2000-2100)")
        return res
    except Exception as err:
        logger.error(f"Invalid date format '{d_str}': {err}")
        raise HTTPException(status_code=400, detail=f"Invalid date format '{d_str}'. Expected YYYY-MM-DD.")


def _get_platform_start(platform: str) -> str:
    return GOOGLE_START_DATE if platform == "google" else META_START_DATE


def _get_mantri_project(db: Session) -> Optional[MisProject]:
    return db.query(MisProject).join(Account).filter(Account.name.ilike("%mantri%"), MisProject.name == "Serenity").first()


def _upsert_snapshot(db: Session, project_id: int, platform: str, snapshot_date: date, leads: float, amount_spent: float):
    existing = db.query(MisDailySnapshot).filter(
        MisDailySnapshot.project_id == project_id,
        MisDailySnapshot.platform == platform,
        MisDailySnapshot.date == snapshot_date,
    ).first()
    cpl = round(amount_spent / leads) if leads else 0
    if existing:
        existing.leads = leads
        existing.amount_spent = amount_spent
        existing.cpl = cpl
    else:
        db.add(MisDailySnapshot(
            project_id=project_id,
            platform=platform,
            date=snapshot_date,
            leads=leads,
            amount_spent=amount_spent,
            cpl=cpl,
        ))
    db.commit()


def _platform_connected(account: Account, platform: str) -> bool:
    if platform == "google":
        return bool(account.google_is_live and account.google_credentials)
    if platform == "meta":
        import os
        return bool(account.meta_is_live and account.meta_credentials) or bool(os.environ.get("META_SYSTEM_USER_TOKEN"))
    return False


def _both_platforms_connected(account: Account) -> bool:
    return _platform_connected(account, "google") and _platform_connected(account, "meta")


def _refresh_platform(db: Session, project: MisProject, platform: str, start: str, end: str, user: User):
    account = db.query(Account).filter(Account.id == project.client_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Mantri client account not found")
    if not _platform_connected(account, platform):
        raise HTTPException(status_code=400, detail=f"{platform} platform is not connected for Mantri")

    connector = get_connector(account, platform, start_date=start, end_date=end)
    if not connector or not connector.is_valid:
        raise HTTPException(status_code=400, detail=f"Unable to initialise {platform} connector for Mantri")

    s_date = _parse_date(start)
    e_date = _parse_date(end)
    rows = connector.fetch_daily_metrics(s_date.isoformat(), e_date.isoformat())
    date_set = {s_date + timedelta(days=i) for i in range((e_date - s_date).days + 1)}

    pulled = 0
    for row in rows:
        d = _parse_date(row["date"])
        _upsert_snapshot(db, project.id, platform, d, float(row.get("leads", 0) or 0), float(row.get("spend", 0) or 0))
        date_set.discard(d)
        pulled += 1

    # Fill missing dates with zeros so the range is fully represented
    for d in sorted(date_set):
        _upsert_snapshot(db, project.id, platform, d, 0.0, 0.0)

    log_activity(
        module="InsightDesk",
        action="Mantri MIS Refresh",
        description=f"Refreshed Mantri MIS {platform} data from {_ordinal_date(s_date)} to {_ordinal_date(e_date)} ({pulled} rows pulled)",
        user_id=user.id,
        user_name=user.full_name or user.email,
        account_id=account.id,
        account_name=account.name,
        entity_type="mis_project",
        entity_id=str(project.id),
        details={"platform": platform, "start": start, "end": end, "pulled": pulled},
        db=db,
    )
    return {"platform": platform, "start": start, "end": end, "pulled": pulled}


def _query_snapshots(db: Session, project_id: int, platforms: List[str], start: str, end: str) -> List[MisDailySnapshot]:
    q = db.query(MisDailySnapshot).filter(
        MisDailySnapshot.project_id == project_id,
        MisDailySnapshot.platform.in_(platforms),
        MisDailySnapshot.date >= _parse_date(start),
        MisDailySnapshot.date <= _parse_date(end),
    ).order_by(MisDailySnapshot.date, MisDailySnapshot.platform)
    return q.all()


# ---------------------------------------------------------------------------
# API models
# ---------------------------------------------------------------------------

class RefreshRequest(BaseModel):
    project_id: int
    platform: str  # meta | google | combined
    start: str
    end: str


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/projects")
def list_mis_projects(db: Session = Depends(get_db), user: User = Depends(get_current_user_required)):
    """List active MIS projects (currently only Mantri Serenity)."""
    projects = db.query(MisProject).filter(MisProject.is_active == True).all()
    return [p.to_dict() for p in projects]


@router.get("/overall")
def overall_summary(
    project_id: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_required),
):
    """Overall summary per platform for the selected end date.

    Start dates are hardcoded per platform:
      - Google: 2026-05-25
      - Meta:   2026-04-17
    The requested `start` is ignored; only `end` is used from the UI.
    """
    project = db.query(MisProject).filter(MisProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    account = db.query(Account).filter(Account.id == project.client_id).first()
    google_connected = _platform_connected(account, "google") if account else False
    meta_connected = _platform_connected(account, "meta") if account else False
    both_connected = google_connected and meta_connected

    effective_starts = []
    rows = []
    end_dt = _parse_date(end)
    for platform in ["google", "meta"]:
        platform_start = _get_platform_start(platform)
        effective_starts.append(_parse_date(platform_start))
        snapshots = _query_snapshots(db, project_id, [platform], platform_start, end_dt.isoformat())
        by_platform: Dict[str, Dict[str, float]] = {}
        for s in snapshots:
            d = by_platform.setdefault(s.platform, {"leads": 0.0, "amount_spent": 0.0})
            d["leads"] += s.leads
            d["amount_spent"] += s.amount_spent
        agg = by_platform.get(platform, {"leads": 0.0, "amount_spent": 0.0})
        leads = agg["leads"]
        spend = agg["amount_spent"]
        cpl = round(spend / leads) if leads else 0
        rows.append({
            "project_name": project.name,
            "platform": platform.title(),
            "amount_spent": spend,
            "leads": leads,
            "cpl": cpl,
        })

    total_leads = sum(r["leads"] for r in rows)
    total_spend = sum(r["amount_spent"] for r in rows)
    total_cpl = round(total_spend / total_leads) if total_leads else 0

    s_date = min(effective_starts)
    e_date = end_dt
    return {
        "title": f"MANTRI MIS – OVERALL ({_ordinal_date(s_date)} – {_ordinal_date(e_date)})",
        "project": project.to_dict(),
        "configured": both_connected,
        "both_connected": both_connected,
        "google_connected": google_connected,
        "meta_connected": meta_connected,
        "start": s_date.isoformat(),
        "end": e_date.isoformat(),
        "rows": rows,
        "total": {
            "project_name": "TOTAL",
            "platform": "",
            "amount_spent": total_spend,
            "leads": total_leads,
            "cpl": total_cpl,
        },
    }


@router.get("/daily")
def daily_report(
    project_id: int = Query(...),
    platform: str = Query(..., enum=["meta", "google", "combined"]),
    start: str = Query(...),
    end: str = Query(...),
    start_from: Optional[str] = Query(None, description="Ignored; platform starts are hardcoded"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_required),
):
    """Daily rows for one platform or combined across platforms.

    Google rows always start from 2026-05-25, Meta rows from 2026-04-17.
    The requested `start`/`start_from` are ignored; only `end` is used from the UI.
    """
    project = db.query(MisProject).filter(MisProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    end_dt = _parse_date(end)

    if platform == "combined":
        platforms = ["meta", "google"]
        effective_starts = [_parse_date(_get_platform_start(p)) for p in platforms]
        effective_start = min(effective_starts).isoformat()
    else:
        effective_start = _get_platform_start(platform)
        platforms = [platform]

    snapshots = _query_snapshots(db, project_id, platforms, effective_start, end_dt.isoformat())

    by_date: Dict[date, Dict[str, Dict[str, float]]] = {}
    for s in snapshots:
        by_date.setdefault(s.date, {}).setdefault(s.platform, {"leads": 0.0, "amount_spent": 0.0})
        by_date[s.date][s.platform]["leads"] += s.leads
        by_date[s.date][s.platform]["amount_spent"] += s.amount_spent

    rows = []
    s_date = _parse_date(effective_start)
    e_date = end_dt
    current = s_date
    while current <= e_date:
        if platform == "combined":
            meta = by_date.get(current, {}).get("meta", {"leads": 0.0, "amount_spent": 0.0})
            google = by_date.get(current, {}).get("google", {"leads": 0.0, "amount_spent": 0.0})
            total_leads = meta["leads"] + google["leads"]
            total_spend = meta["amount_spent"] + google["amount_spent"]
            rows.append({
                "date": current.isoformat(),
                "display_date": _ordinal_date(current),
                "meta_leads": meta["leads"],
                "meta_cpl": round(meta["amount_spent"] / meta["leads"]) if meta["leads"] else 0,
                "meta_amount_spent": meta["amount_spent"],
                "google_leads": google["leads"],
                "google_cpl": round(google["amount_spent"] / google["leads"]) if google["leads"] else 0,
                "google_amount_spent": google["amount_spent"],
                "leads": total_leads,
                "cpl": round(total_spend / total_leads) if total_leads else 0,
                "amount_spent": total_spend,
            })
        else:
            agg = by_date.get(current, {}).get(platform, {"leads": 0.0, "amount_spent": 0.0})
            leads = agg["leads"]
            spend = agg["amount_spent"]
            rows.append({
                "date": current.isoformat(),
                "display_date": _ordinal_date(current),
                "leads": leads,
                "cpl": round(spend / leads) if leads else 0,
                "amount_spent": spend,
            })
        current += timedelta(days=1)

    total_leads = sum(r["leads"] for r in rows)
    total_spend = sum(r["amount_spent"] for r in rows)
    total_cpl = round(total_spend / total_leads) if total_leads else 0

    title_prefix = platform.upper() if platform != "combined" else "META ADS + GOOGLE COMBINED"
    if platform == "combined":
        title = f"{title_prefix} (From {_ordinal_date(s_date)})"
    else:
        title = f"{title_prefix} ({_ordinal_date(s_date)} – {_ordinal_date(e_date)})"

    account = db.query(Account).filter(Account.id == project.client_id).first()
    google_connected = _platform_connected(account, "google") if account else False
    meta_connected = _platform_connected(account, "meta") if account else False
    both_connected = google_connected and meta_connected

    return {
        "title": title,
        "project": project.to_dict(),
        "configured": both_connected,
        "both_connected": both_connected,
        "google_connected": google_connected,
        "meta_connected": meta_connected,
        "platform": platform,
        "start": effective_start,
        "end": e_date.isoformat(),
        "rows": rows,
        "total": {
            "date": "",
            "display_date": "TOTAL",
            "leads": total_leads,
            "cpl": total_cpl,
            "amount_spent": total_spend,
            "meta_leads": sum(r.get("meta_leads", 0) for r in rows),
            "meta_amount_spent": sum(r.get("meta_amount_spent", 0) for r in rows),
            "google_leads": sum(r.get("google_leads", 0) for r in rows),
            "google_amount_spent": sum(r.get("google_amount_spent", 0) for r in rows),
        },
    }


@router.post("/refresh")
def refresh_data(
    req: RefreshRequest,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_required),
):
    """Pull live data from ad platforms and upsert into mis_daily_snapshots."""
    project = db.query(MisProject).filter(MisProject.id == req.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    start_dt = _parse_date(req.start)
    end_dt = _parse_date(req.end)
    valid_start = start_dt.isoformat()
    valid_end = end_dt.isoformat()

    platforms = ["meta", "google"] if req.platform == "combined" else [req.platform]
    results = []
    skipped = []
    for platform in platforms:
        account = db.query(Account).filter(Account.id == project.client_id).first()
        if not account:
            raise HTTPException(status_code=404, detail="Mantri client account not found")
        if not _platform_connected(account, platform):
            skipped.append(platform)
            logger.info(f"Mantri MIS refresh: skipping {platform} (not connected)")
            continue
        try:
            results.append(_refresh_platform(db, project, platform, valid_start, valid_end, user))
        except HTTPException as e:
            skipped.append(platform)
            logger.warning(f"Mantri MIS refresh: {platform} failed: {e.detail}")
        except Exception as e:
            skipped.append(platform)
            logger.error(f"Mantri MIS refresh: {platform} failed unexpectedly: {e}")

    if not results and skipped:
        skipped_str = ", ".join(skipped)
        raise HTTPException(
            status_code=400,
            detail=f"No platforms could be refreshed. Not connected or failed: {skipped_str}. "
                   f"Please check account connections in AdPulse → Manage Accounts."
        )

    return {
        "status": "success",
        "project_id": project.id,
        "results": results,
        "skipped": skipped,
    }


@router.get("/export")
def export_report(
    type: str = Query(..., enum=["pdf", "xlsx"]),
    report: str = Query(..., enum=["overall", "meta", "google", "combined"]),
    project_id: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    start_from: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user_required),
):
    """Export an Mantri MIS report as PDF or Excel."""
    platform_map = {"overall": "combined", "meta": "meta", "google": "google", "combined": "combined"}
    platform = platform_map[report]

    if report == "overall":
        payload = overall_summary(project_id=project_id, start=start, end=end, db=db, user=user)
        rows = payload["rows"]
        total = payload["total"]
        title = payload["title"]
        headers = ["Project", "Platform", "Amount Spent", "Leads", "CPL"]
        data_rows = [[r["project_name"], r["platform"], _fmt_inr(r["amount_spent"]), str(int(r["leads"])), _fmt_inr(r["cpl"])] for r in rows]
        data_rows.append(["TOTAL", "", _fmt_inr(total["amount_spent"]), str(int(total["leads"])), _fmt_inr(total["cpl"])])
        file_title = "Overall Summary"
        s_date = _parse_date(payload["start"])
        e_date = _parse_date(payload["end"])
    else:
        payload = daily_report(project_id=project_id, platform=platform, start=start, end=end, start_from=start_from, db=db, user=user)
        rows = payload["rows"]
        total = payload["total"]
        title = payload["title"]
        if platform == "combined":
            headers = ["Day", "Leads", "CPL", "Amount spent"]
            data_rows = [[r["display_date"], str(int(r["leads"])), _fmt_inr(r["cpl"]), _fmt_inr(r["amount_spent"])] for r in rows]
            data_rows.append(["TOTAL", str(int(total["leads"])), _fmt_inr(total["cpl"]), _fmt_inr(total["amount_spent"])])
        else:
            headers = ["Date", "Leads", "CPL", "Amount Spent"]
            data_rows = [[r["display_date"], str(int(r["leads"])), _fmt_inr(r["cpl"]), _fmt_inr(r["amount_spent"])] for r in rows]
            data_rows.append(["TOTAL", str(int(total["leads"])), _fmt_inr(total["cpl"]), _fmt_inr(total["amount_spent"])])
        file_title = platform.title() + " Daily"

    if report == "combined":
        s_date = date.fromisoformat(payload["start"])
        e_date = date.fromisoformat(end)
    else:
        s_date = date.fromisoformat(_get_platform_start(platform))
        e_date = date.fromisoformat(end)

    # Activity log
    account = db.query(Account).filter(Account.id == payload["project"]["client_id"]).first()
    log_activity(
        module="InsightDesk",
        action="Mantri MIS Export",
        description=f"Exported Mantri MIS {file_title} as {type.upper()} from {_ordinal_date(s_date)} to {_ordinal_date(e_date)}",
        user_id=user.id,
        user_name=user.full_name or user.email,
        account_id=account.id if account else None,
        account_name=account.name if account else None,
        entity_type="mis_project",
        entity_id=str(project_id),
        details={"report": report, "type": type, "start": s_date.isoformat(), "end": e_date.isoformat()},
        db=db,
    )

    if type == "xlsx":
        return _export_xlsx(title, headers, data_rows, report)
    return _export_pdf(title, headers, data_rows, report)


def _export_xlsx(title: str, headers: List[str], rows: List[List[str]], report: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = report[:31]
    ws.append([title])
    ws.append(headers)
    for r in rows:
        ws.append(r)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    title_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for cell in ws[1]:
        cell.font = title_font
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        is_total = str(row[0].value or "").upper() == "TOTAL"
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if cell.column != 1 else "left", vertical="center")
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid")

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Mantri_MIS_{report}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_pdf(title: str, headers: List[str], rows: List[List[str]], report: str):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    warm_slate = colors.HexColor("#64748b")
    dark_red = colors.HexColor("#991B1B")
    light_red = colors.HexColor("#FDE2E2")

    title_style = ParagraphStyle(
        "MantriMISTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        alignment=1,
        textColor=dark_red,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        "MantriMISSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        alignment=1,
        textColor=warm_slate,
        spaceAfter=12,
    )

    elements = [Paragraph("Mantri MIS Report", title_style), Paragraph(title, subtitle_style), Spacer(1, 0.3 * cm)]

    table_data = [headers]
    table_data.extend(rows)

    col_count = len(headers)
    available_width = A4[0] - 3 * cm
    col_width = available_width / col_count
    t = Table(table_data, colWidths=[col_width] * col_count)

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), dark_red),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
    ]
    # Highlight total row
    for i, row in enumerate(rows, start=1):
        if str(row[0] or "").upper() == "TOTAL":
            style_commands.append(("BACKGROUND", (0, i), (-1, i), light_red))
            style_commands.append(("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"))
            style_commands.append(("TEXTCOLOR", (0, i), (-1, i), dark_red))

    t.setStyle(TableStyle(style_commands))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    filename = f"Mantri_MIS_{report}_{date.today().isoformat()}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
