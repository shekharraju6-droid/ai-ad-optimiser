"""Push local Excel report to a new Google Sheet for Shyam Steel.

Uses a stored Gmail-user refresh token (Google Sheets OAuth) to create a
new spreadsheet and populate it with the first 3 tabs of
`data/shym_steel_report.xlsx`.
"""
import json
import logging
from typing import Any, Dict, List, Optional

from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials as UserCredentials
from google.auth.transport.requests import Request
from openpyxl import load_workbook

from backend.services.config import load_config
from backend.services.crypto import decrypt
from backend.db.models import Account

logger = logging.getLogger("AdOptima")

EXCEL_PATH = "data/shym_steel_report.xlsx"


def _get_sheets_credentials(refresh_token: str) -> UserCredentials:
    cfg = load_config()
    client_id = cfg.get("gmail_client_id") or cfg.get("google_client_id", "")
    client_secret = cfg.get("gmail_client_secret") or cfg.get("google_client_secret", "")
    if not client_id or not client_secret or not refresh_token:
        raise RuntimeError("Missing Google Sheets OAuth credentials")
    creds = UserCredentials(
        None,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    creds.refresh(Request())
    return creds


def _read_excel_sheets() -> Dict[str, List[List[Any]]]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    data: Dict[str, List[List[Any]]] = {}
    # Only first 3 tabs
    for name in wb.sheetnames[:3]:
        ws = wb[name]
        rows: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            # Convert datetime objects to strings; None stays None
            cleaned = [
                v.isoformat() if hasattr(v, "isoformat") else v
                for v in row
            ]
            rows.append(cleaned)
        data[name] = rows
    wb.close()
    return data


def push_excel_to_new_google_sheet(account: Account) -> Dict[str, Any]:
    """Create a new Google Sheet from Shyam Steel Excel and return its URL."""
    # Load refresh token
    refresh_token = None
    if account.crm_credentials:
        try:
            raw = decrypt(account.crm_credentials)
            cfg = json.loads(raw)
            refresh_token = cfg.get("sheets_refresh_token")
        except Exception as e:
            logger.warning(f"Could not read sheets refresh token: {e}")
    if not refresh_token:
        raise RuntimeError("Google Sheets not linked. Click 'Link Google Sheet' first.")

    creds = _get_sheets_credentials(refresh_token)
    service = build("sheets", "v4", credentials=creds)

    # Create new spreadsheet
    spreadsheet = service.spreadsheets().create(
        body={"properties": {"title": f"Shyam Steel - Marketing Report ({account.name})"}}
    ).execute()
    spreadsheet_id = spreadsheet["spreadsheetId"]

    excel_data = _read_excel_sheets()

    # Rename default sheet to first tab name and add remaining sheets
    sheet_titles = list(excel_data.keys())
    if not sheet_titles:
        raise RuntimeError("No sheets found in Excel file")

    # First sheet is already "Sheet1"; rename it to first Excel tab
    requests = [
        {"updateSheetProperties": {"properties": {"sheetId": 0, "title": sheet_titles[0]}, "fields": "title"}}
    ]
    for title in sheet_titles[1:]:
        requests.append({"addSheet": {"properties": {"title": title}}})

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id, body={"requests": requests}
    ).execute()

    # Get sheet IDs
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheet_id_by_title = {
        s["properties"]["title"]: s["properties"]["sheetId"]
        for s in spreadsheet["sheets"]
    }

    # Write data to each sheet
    for title, rows in excel_data.items():
        if not rows:
            continue
        # Clear grid and resize to data
        num_rows = len(rows)
        num_cols = max(len(r) for r in rows)
        body = {"values": rows}
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{title}!A1",
            valueInputOption="USER_ENTERED",
            body=body,
        ).execute()

        # Auto-resize columns (basic)
        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={
                "requests": [
                    {
                        "autoResizeDimensions": {
                            "dimensions": {
                                "sheetId": sheet_id_by_title[title],
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": num_cols,
                            }
                        }
                    }
                ]
            },
        ).execute()

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    return {
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_url": url,
        "message": "New Google Sheet created from Excel report.",
        "tabs": sheet_titles,
    }
